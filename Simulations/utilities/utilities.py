# coding=utf-8
import pandas as pd
import numpy as np
import math
import statistics
from scipy.stats import poisson, norm, chi2, gamma, uniform
import scipy.optimize as opt
from openpyxl import load_workbook


def write_to_exist_excel1(data, fileName, sheetName):
    """
    Write data to Excel file

    Args:
        data: data to write
        fileName: name of Excel file
        sheetName: name of sheet to write
    """
    df_old = pd.DataFrame(pd.read_excel(fileName, sheet_name=sheetName))
    row_old = df_old.shape[0]

    df = data

    book = load_workbook(fileName)
    writer = pd.ExcelWriter(fileName, engine='openpyxl')
    writer._book = book
    writer._sheets = dict((ws.title, ws) for ws in book.worksheets)

    df.to_excel(writer, sheet_name=sheetName, startrow=row_old + 1, index=False, header=False)

    writer._save()


def write_to_exist_excel2(data, fileName, sheetName):
    """
        Write type(list[dict]) data to Excel file

        Args:
            data: data to write
            fileName: name of Excel file
            sheetName: name of sheet to write
    """
    df_old = pd.DataFrame(pd.read_excel(fileName, sheet_name=sheetName))
    row_old = df_old.shape[0]

    df = data

    book = load_workbook(fileName)
    writer = pd.ExcelWriter(fileName, engine='openpyxl')
    writer._book = book
    writer._sheets = dict((ws.title, ws) for ws in book.worksheets)

    df.to_excel(writer, sheet_name=sheetName, startrow=row_old + 1, index=False, header=False)

    writer._save()


def normfun(x,mu,sigma):
    """
        Compute the pdf of a normal distribution

        Args:
            x: data
            mu: mean
            sigma: standard deviation

        Returns:
            pdf of normal distribution
    """
    return norm.pdf(x,loc=mu,scale=sigma)

def poisson_data(mu):
    """
        Generate Poisson data given the expectations

        Args:
            mu: vector of means

        Returns:
            Poisson data
    """
    return np.random.poisson(mu)


def nb_data(r,mu):
    """
        Generate Negative Binomial data given the parameters

        Args:
            r: number of trials
            mu: vector of means

        Returns:
            Negative Binomial data
    """
    prob = r/(r+mu)
    return np.random.negative_binomial(r,prob)

def p_value_norm(x,mu,sigma):
    """
    Compute the one-sided and two-sided p-value of a normal distribution

    Args:
        x: data
        mu: vector of means
        sigma: standard deviation

    Returns:
        norm.sf(z): one-sided p-value
        2*min(norm.sf(z),1-norm.sf(z)): two-sided p-value
    """
    z=(x-mu)/sigma
    return norm.sf(z), 2*min(norm.sf(z),1-norm.sf(z))

def generate_s_exp(n, beta):
    """
    Generate expected counts s with a exponential spectral

    Args:
        n: number of bins
        beta: parameter

    Returns:
        expected counts s with a exponential spectral
    """
    theta1, theta2 = beta
    i_over_n = np.arange(n) / n
    return theta1 * np.exp(-theta2 * i_over_n)

def generate_s_constant(n,mu):
    """
    Generate expected counts s with a constant spectral

    Args:
        n: number of bins
        mu: parameter

    Returns:
        expected counts s with a constant spectral
    """
    return np.repeat(mu,n)

def generate_s_powerlaw(n, beta):
    """
    Generate expected counts s with a powerlaw spectral

    Args:
        n: number of bins
        beta: parameter

    Returns:
        expected counts s with a powerlaw spectral
    """
    mu, slope = beta
    i_values = np.arange(1, n+1)
    return mu * (1 + i_values / n) ** (-slope)


def generate_s_broken_powerlaw(n, beta, loc=0.5, slope2=0.5):
    """
    Generate expected counts s with a broken powerlaw spectral

    Args:
        n: number of bins
        beta: parameter
        loc: location of changing point
        slope2: slope of second powerlaw spectral

    Returns:
        expected counts s with a broken powerlaw spectral
    """
    mu, slope1 = beta
    breakpoint = int(n * loc)
    if breakpoint == 0:
        return generate_s_powerlaw(n, np.array([mu, slope2]))

    # Before loc
    arr = np.empty(n)
    i_part1 = np.arange(1, breakpoint + 1)
    arr[:breakpoint] = mu * (1 + i_part1 / n) ** (-slope1)

    # After loc
    i_part2 = np.arange(1, n - breakpoint + 1)
    decay_factor = (1 + i_part2 / (n - breakpoint)) ** (-slope2)
    arr[breakpoint:] = arr[breakpoint - 1] * decay_factor

    return arr

def generate_s_log_norm(n,mu,sigma):
    """
    Generate expected counts s with a log-normal spectral

    Args:
        n: number of bins
        mu, sigma: parameter

    Returns:
        expected counts s with a log-normal spectral
    """
    return np.exp(np.random.normal(np.log(mu),sigma,n))

def generate_s_gamma(n,alpha,beta):
    """
    Generate expected counts s with a Gamma spectral

    Args:
        n: number of bins
        alpha, beta: parameter

    Returns:
        expected counts s with a Gamma spectral
    """
    return gamma.rvs(alpha,loc=0,scale=1/beta,size=n)

def generate_s_unif(n,mu):
    """
    Generate expected counts s with a uniform spectral

    Args:
        n: number of bins
        mu: parameter

    Returns:
        expected counts s with a uniform spectral
    """
    return uniform.rvs(0,2*mu,size=n)

def generate_s(n,beta,snull,loc=0.5,strength=0.5,width=5):
    """
    Generate expected counts s under the null H0

    Args:
        n: number of bins
        beta, loc, strength, width: parameter
        snull: null H0

    Returns:
        expected counts s under the null H0
    """
    if snull=='exp':
        return generate_s_exp(n,beta)
    elif snull=='powerlaw':
        return generate_s_powerlaw(n,beta)
    elif snull=='constant':
        return generate_s_constant(n,beta[0])
    elif snull=='brokenpowerlaw':
        return generate_s_broken_powerlaw(n,beta,loc=loc,slope2=strength)
    else:
        raise ValueError("Warning: Invalid type of distribution!")


def spectral_line(s, loc=0.75, strength=10, width=3):
    """
    Generate expected counts s with a emission/ absorption line

    Args:
        s: original expected counts without the line
        loc: location of spectral line
        strength: line strength
        width: width of spectral line

    Returns:
        expected counts s with a emission/ absorption line
    """
    position = int(len(s) * loc)
    end = min(position + width, len(s))
    s[position:end] = strength
    return s

def generate_s_true(n,beta,strue,snull,loc=0.5,strength=10,width=5):
    """
    Generate expected counts s under the true model

    Args:
        n: number of bins
        beta, loc, strength, width: parameter
        snull: true model

    Returns:
        expected counts s under the true model
    """
    if strue=='exp':
        return generate_s_exp(n,beta)
    elif strue=='powerlaw':
        return generate_s_powerlaw(n,beta)
    elif strue=='constant':
        return generate_s_constant(n,beta[0])
    elif strue=='lognorm':
        return generate_s_log_norm(n,beta[0],beta[1])
    elif strue=='gamma':
        return generate_s_gamma(n,beta[0],beta[1])
    elif strue=='brokenpowerlaw':
        return generate_s_broken_powerlaw(n,beta,loc=loc,slope2=strength)
    elif strue=='unif':
        return generate_s_unif(n,beta[0])
    elif strue=='spectral_line':
        return spectral_line(generate_s(n,beta,snull),loc=loc,strength=strength,width=width)
    else:
        raise ValueError("Warning: Invalid type of distribution!")

def empirical_bounds(x,snull,epsilon=1e-5):
    """
    Compute the empirical bounds of the expected counts for optimization

    Args:
        x: data
        snull: null H0
        epsilon: small stablizer

    Returns:
        bound: empirical bounds
    """
    bound = [[epsilon, 2*max(x)]] if snull == 'constant' else [[epsilon, 2*max(x)],
                                                             [-2*math.log((max(x) + 2 * epsilon) / (min(x) + epsilon), 2),
                                                              2*math.log((max(x) + 2 * epsilon) / (min(x) + epsilon), 2)]]
    return bound

def Cashstat(x,s):
    """
    Compute the Cash statistics of the data

    Args:
        x: data
        s: expected counts

    Returns:
        C: Cash statistics
    """
    C=0.
    for j in range(len(x)):
        if x[j] <1e-5:
            C+=s[j]
        else:
            C +=s[j] - x[j] * np.log(s[j] / x[j]) - x[j]
    C=C*2
    return C