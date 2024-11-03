import pandas as pd
import numpy as np
import math
import matplotlib.pyplot as plt
import random
import statistics
from scipy.stats import poisson
from scipy.stats import chi2
from scipy.stats import nbinom
import scipy.optimize as opt
import scipy
import time
import pandas as pd
from openpyxl import load_workbook
import os
np.set_printoptions(threshold=np.inf)

def write_to_exist_excel1(data,fileName, sheetName):
    df_old = pd.DataFrame(pd.read_excel(fileName, sheet_name=sheetName))  # 读取原数据文件和表
    row_old = df_old.shape[0]  # 获取原数据的行数

    df = data

    book = load_workbook(fileName)
    writer = pd.ExcelWriter(fileName, engine='openpyxl')
    writer._book = book
    writer._sheets = dict((ws.title, ws) for ws in book.worksheets)

    # 将data_added数据写入Excel中
    df.to_excel(writer, sheet_name=sheetName, startrow=row_old + 1, index=False, header=False)

    writer._save()  # 保存


# 将list[dict]类型的数据追加写入到现有的Excel中
def write_to_exist_excel2(data,fileName, sheetName):
    df_old = pd.DataFrame(pd.read_excel(fileName, sheet_name=sheetName))  # 读取原数据文件和表
    row_old = df_old.shape[0]  # 获取原数据的行数

    df = data

    book = load_workbook(fileName)
    writer = pd.ExcelWriter(fileName, engine='openpyxl')
    writer._book = book
    writer._sheets = dict((ws.title, ws) for ws in book.worksheets)

    # 将data_added数据写入Excel中
    df.to_excel(writer, sheet_name=sheetName, startrow=row_old + 1, index=False, header=False)

    writer._save()  # 保存


def normfun(x,mu,sigma):
    pdf=np.exp(-((x-mu)**2)/(2*sigma**2))/(sigma*np.sqrt(2*np.pi))
    return pdf

def poisson_data(mu):
    arr=[0 for x in range(len(mu))]
    for i in range(len(mu)):
        arr[i]=poisson.rvs(mu=mu[i],size=1)[0]
    return arr

def poisson_data_constant(mu,n):
    return poisson.rvs(mu=mu,size=n).tolist()

def nb_data(r,mu):
    arr = [0. for x in range(len(mu))]
    prob=[0. for x in range(len(mu))]
    for i in range(len(mu)):
        prob[i]=r/(r+mu[i])
    for i in range(len(mu)):
        arr[i] = nbinom.rvs(r,prob[i],size=1)[0]
    return arr

def p_value_norm(x,mu,sigma):
    z=(x-mu)/sigma
    return scipy.stats.norm.sf(z)

def p_value_chi(x,df):
    return scipy.stats.chi2.sf(x,df)

def generate_s_exp(n,theta1,theta2):
    arr=[0 for x in range(n)]
    for i in range(n):
        arr[i]=theta1*math.exp(-theta2*i/n)
    return arr

def generate_s_constant(n,mu):
    return [mu for i in range(n)]

def generate_s_powerlaw(n,theta1,theta2):
    arr=[0 for x in range(n)]
    for i in range(n):
        arr[i]=theta1*((1+(i+1)/n)**theta2)
    return arr

def generate_s_norm(n,mu,sigma):
    return np.random.normal(mu,sigma,n)

def generate_s_gamma(n,alpha,beta):
    return scipy.stats.gamma.rvs(alpha,loc=0,scale=1/beta,size=n)

def poisson_dis(mu,i):
    return mu**i/math.factorial(i)*math.e**(-mu)

#def Sigma_diag(Q,n):
    #return kapa_12-kapa_11*Q*kapa_03*n

def con_expectation(beta,n,kapa_1,kapa_2,kapa_11,kapa_12,kapa_03,X,I):
    mu=beta[0]
    #V = np.diag([mu for i in range(n)])
    #Q=X*X.T/(mu*n)
    #sigma=Sigma_diag(1/(mu*n),n)
    sigma = kapa_12-kapa_11*kapa_03/mu
    #Sigma=np.diag([sigma for i in range(n)])
    E=-0.5*sigma/mu
    E+=kapa_1*n
    return E

def uncon_expectation(beta,n,kapa_1,kapa_2,kapa_11,kapa_12,kapa_03,X,I):
    #mu=beta[0]
    #V = np.diag([mu for i in range(n)])
    #Q=X*X.T/(mu*n)
    #sigma=Sigma_diag(1/(mu*n),n)
    #sigma = kapa_12-kapa_11*kapa_03/mu
    #Sigma=np.diag([sigma for i in range(n)])
    #E=-0.5*sigma/mu
    E=kapa_1*n
    return E

def Var(beta,n,kapa_1,kapa_2,kapa_11,kapa_12,kapa_03,X,I):
    mu=beta[0]
    #V = np.diag([mu for i in range(n)])
    #k_11=np.mat([kapa_11 for i in range(n)]).T
    #var=(-k_11.T*X*X.T*k_11)[0,0]/(mu*n)
    var=-kapa_11**2*n/mu
    var+=kapa_2*n
    return var

def kapa1(mu,max):
    x = 0
    for k in range(max):
        if k == 0:
            x -= 2 * (k - mu) * poisson_dis(mu, k)
        else:
            x += 2 * (k * math.log(k / mu, math.e) - k + mu) * poisson_dis(mu, k)
    return x

def kapa2(mu,max):
    x = 0
    k_1=kapa1(mu,max)
    for k in range(max):
        if k == 0:
            x += (-2 * (k - mu) - k_1) ** 2 * poisson_dis(mu, k)
        else:
            x += (2 * (k * math.log(k / mu, math.e) - k + mu) - k_1) ** 2 * poisson_dis(mu, k)
    return x

def kapa11(mu,max):
    x = 0
    k_1 = kapa1(mu, max)
    for k in range(max):
        if k == 0:
            x += (-2 * (k - mu) - k_1) * (k - mu) * poisson_dis(mu, k)
        else:
            x += (2 * (k * math.log(k / mu, math.e) - k + mu) - k_1) * (k - mu) * poisson_dis(mu, k)
    return x

def kapa12(mu,max):
    x = 0
    k_1 = kapa1(mu, max)
    for k in range(max):
        if k == 0:
            x += (-2 * (k - mu) - k_1) * (k - mu) ** 2 * poisson_dis(mu, k)
        else:
            x += (2 * (k * math.log(k / mu, math.e) - k + mu) - k_1) * (k - mu) ** 2 * poisson_dis(mu,k)
    return x

def kapa03(mu):
    return mu

def con_theory_test(x,beta,n,X,I,max=30):
    mu_hat=beta[0]

    kapa_1=kapa1(mu_hat,max)

    kapa_2 =kapa2(mu_hat,max)

    kapa_11 = kapa11(mu_hat,max)

    kapa_12 = kapa12(mu_hat,max)

    kapa_03 = kapa03(mu_hat)

    return p_value_norm(x,con_expectation(beta,n,kapa_1,kapa_2,kapa_11,kapa_12,kapa_03,X,I),math.sqrt(Var(beta,n,kapa_1,kapa_2,kapa_11,kapa_12,kapa_03,X,I)))


def uncon_theory_test(x,beta,n,X,I,max=30):
    mu_hat=beta[0]

    kapa_1=kapa1(mu_hat,max)

    kapa_2 =kapa2(mu_hat,max)

    kapa_11 = kapa11(mu_hat,max)

    kapa_12 = kapa12(mu_hat,max)

    kapa_03 = kapa03(mu_hat)

    return p_value_norm(x,uncon_expectation(beta,n,kapa_1,kapa_2,kapa_11,kapa_12,kapa_03,X,I),math.sqrt(Var(beta,n,kapa_1,kapa_2,kapa_11,kapa_12,kapa_03,X,I)))

'''
def bootstrap_test(Cmin,beta,n,B=1000):
    mu_hat=beta[0]
    C = [0 for x in range(B)]
    for i in range(B):
        x = poisson_data_constant(mu_hat,n)
        mu_mle=np.mean(x)
        r = generate_s_constant(n, mu_mle)
        for j in range(n):
            if x[j] == 0:
                C[i] +=r[j]
            else:
                C[i] +=r[j] - x[j] * math.log(r[j]/x[j], math.e) - x[j]
        C[i]=2*C[i] # 2 is here
    k=0
    for i in range(B):
        if C[i]>=Cmin:
            k+=1
    return k/B

def bootstrap_CAN(Cmin,beta,n,B=1000):
    mu_hat=beta[0]
    C = [0 for x in range(B)]
    for i in range(B):
        x = poisson_data_constant(mu_hat,n)
        mu_mle=np.mean(x)
        r = generate_s_constant(n, mu_mle)
        for j in range(n):
            if x[j] == 0:
                C[i] +=r[j]
            else:
                C[i] +=r[j] - x[j] * math.log(r[j]/x[j], math.e) - x[j]
        C[i]=2*C[i] # 2 is here
    return p_value_norm(Cmin, statistics.mean(C), statistics.stdev(C))

def double_pb(Cmin,beta,n,B1=1000,B2=1000):
    mu_hat=beta[0]
    C=[0 for x in range(B1)]
    pvalue=[0 for x in range(B1)]
    for i in range(B1):
        x = poisson_data_constant(mu_hat, n)
        mu_mle = np.mean(x)
        r = generate_s_constant(n, mu_mle)
        for j in range(n):
            if x[j] == 0:
                C[i] += r[j]
            else:
                C[i] += r[j] - x[j] * math.log(r[j] / x[j], math.e) - x[j]
        C[i] = 2 * C[i]  # 2 is here
        pvalue[i]=bootstrap_test(C[i],[mu_mle],n,B2)
    k = 0
    for i in range(B1):
        if C[i] >= Cmin:
            k += 1
    p=k/B1
    l=0
    for i in range(B1):
        if pvalue[i]<=p:
            l+=1
    return l/B1

def n_boostrap(Cmin,data,B=1000):
    n = len(data)
    C = [0 for x in range(B)]
    for i in range(B):
        x = resample(data)
        mu_mle = np.mean(x)
        r = generate_s_constant(n, mu_mle)
        for j in range(n):
            if x[j] == 0:
                C[i] += r[j]
            else:
                C[i] += r[j] - x[j] * math.log(r[j] / x[j], math.e) - x[j]
        C[i] = 2 * C[i]  # 2 is here
    k = 0
    for i in range(B):
        if C[i] >= Cmin:
            k += 1
    return k / B

def double_npb(Cmin,data,B1=1000,B2=1000):
    n=len(data)
    C=[0 for x in range(B1)]
    pvalue=[0 for x in range(B1)]
    for i in range(B1):
        x =resample(data)
        mu_mle = np.mean(x)
        r = generate_s_constant(n, mu_mle)
        for j in range(n):
            if x[j] == 0:
                C[i] += r[j]
            else:
                C[i] += r[j] - x[j] * math.log(r[j] / x[j], math.e) - x[j]
        C[i] = 2 * C[i]  # 2 is here
        pvalue[i]=n_boostrap(C[i],x,B2)
    k = 0
    for i in range(B1):
        if C[i] >= Cmin:
            k += 1
    p=k/B1
    l=0
    for i in range(B1):
        if pvalue[i]<=p:
            l+=1
    return l/B1

def resample(data):
    n=len(data)
    return [data[random.randint(0,n-1)] for j in range(n)]

def bootstrap_biase(Cmin,beta,n,B1=1000,B2=1000):
    mu_hat=beta[0]
    mu_tilde=[0 for x in range(B1)]
    for i in range(B1):
        x = poisson_data_constant(mu_hat, n)
        mu_tilde[i]=np.mean(x)
    mu_adj=2*mu_hat-np.mean(mu_tilde)
    return bootstrap_test(Cmin,[mu_adj],n,B2)
'''

def test(n,B,B1,B2,beta,max,X,I,iters):
    reject1 = [0 for y in range(iters)]  # uncon
    reject2 = [0 for y in range(iters)]  # con

    for l in range(iters):
        s = generate_s_gamma(n,beta[0],math.sqrt(beta[0]))
        x=poisson_data(s)
        if x == [0 for i in range(n)]:
            continue
        mu_hat = np.mean(x)
        r = generate_s_constant(n, mu_hat)
        Cmin = 0.
        for j in range(n):
            if x[j] == 0:
                Cmin += r[j]
            else:
                Cmin += r[j] - x[j] * math.log(r[j] / x[j], math.e) - x[j]
        Cmin = 2 * Cmin

        if uncon_theory_test(Cmin, [mu_hat], n,X,I,max) < 0.05:
            reject1[l] = 1

        if con_theory_test(Cmin, [mu_hat], n,X,I,max) < 0.05:
            reject2[l] = 1

    print(beta,np.mean(reject1), np.mean(reject2))
    return np.mean(reject1), np.mean(reject2)




B = 100
B1 = 100
B2 = 100
sqrtalpha=[0.1*(i+1) for i in range(100)]
ns=[5,10,15,25,35,50,75,100]
max = 50

power1=np.zeros((1,len(sqrtalpha)))  #uncon
power2=np.zeros((1,len(sqrtalpha)))  #con

n=ns[7]
X = np.mat([[1. for x in range(n)]]).T
I = np.mat([1.]).T

i=0
for alpha in sqrtalpha:
    iters = 10000
    power1[0,i],power2[0,i]=test(n, B, B1, B2, [alpha**2], max, X, I, iters)
    i=i+1

filePath='power_model_C.xlsx'
sheetName1='uncondata'
sheetName2='condata'
result=pd.DataFrame(power1)
if not os.path.exists(filePath):
    result.to_excel(filePath,sheet_name=sheetName1,index=False)
else:
    write_to_exist_excel1(result,filePath,sheetName1)

result=pd.DataFrame(power2)
if not os.path.exists(filePath):
    result.to_excel(filePath,sheet_name=sheetName2,index=False)
else:
    write_to_exist_excel1(result,filePath,sheetName2)

'''
n=ns[7]
X = np.mat([[1. for x in range(n)]]).T
I = np.mat([1.]).T

i=0
for alpha in sqrtalpha:
    iters = 10000
    power1[0,i],power2[0,i]=test(n, B, B1, B2, [alpha**2], max, X, I, iters)
    i=i+1

filePath='power_model_C.xlsx'
sheetName1='uncondata'
sheetName2='condata'
result=pd.DataFrame(power1)
if not os.path.exists(filePath):
    result.to_excel(filePath,sheet_name=sheetName1,index=False)
else:
    write_to_exist_excel1(result,filePath,sheetName1)

result=pd.DataFrame(power2)
if not os.path.exists(filePath):
    result.to_excel(filePath,sheet_name=sheetName2,index=False)
else:
    write_to_exist_excel1(result,filePath,sheetName2)
'''