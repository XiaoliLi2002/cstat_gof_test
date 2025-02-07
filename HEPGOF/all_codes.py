import pandas as pd
import numpy as np
import math
import statistics
from scipy.stats import poisson, norm, chi2, gamma, uniform
import scipy.optimize as opt
from openpyxl import load_workbook

def empirical_bounds(x,snull,epsilon=1e-5):
    bound = [[epsilon, max(x)]] if snull == 'constant' else [[epsilon, max(x)],
                                                             [-math.log((max(x) + 2 * epsilon) / (min(x) + epsilon), 2),
                                                              math.log((max(x) + 2 * epsilon) / (min(x) + epsilon), 2)]]
    return bound

def write_to_exist_excel1(data,fileName, sheetName):
    df_old = pd.DataFrame(pd.read_excel(fileName, sheet_name=sheetName))  # 读取原数据文件和表
    row_old = df_old.shape[0]  # 获取原数据的行数

    df = data

    book = load_workbook(fileName)
    writer = pd.ExcelWriter(fileName, engine='openpyxl')
    writer._book = book
    writer._sheets = dict((ws.title, ws) for ws in book.worksheets)

    # 将data_added数据写入Excel中
    df.to_excel(writer, sheet_name=sheetName, startrow=row_old + 1, index=False, header=False)

    writer._save()  # 保存


# 将list[dict]类型的数据追加写入到现有的Excel中
def write_to_exist_excel2(data,fileName, sheetName):
    df_old = pd.DataFrame(pd.read_excel(fileName, sheet_name=sheetName))  # 读取原数据文件和表
    row_old = df_old.shape[0]  # 获取原数据的行数

    df = data

    book = load_workbook(fileName)
    writer = pd.ExcelWriter(fileName, engine='openpyxl')
    writer._book = book
    writer._sheets = dict((ws.title, ws) for ws in book.worksheets)

    # 将data_added数据写入Excel中
    df.to_excel(writer, sheet_name=sheetName, startrow=row_old + 1, index=False, header=False)

    writer._save()  # 保存


def normfun(x,mu,sigma):
    pdf=np.exp(-((x-mu)**2)/(2*sigma**2))/(sigma*np.sqrt(2*np.pi))
    return pdf

def poisson_data(mu):
    return np.random.poisson(mu)


def nb_data(r,mu):
    prob = r/(r+mu)
    return np.random.negative_binomial(r,prob)

def p_value_norm(x,mu,sigma):
    z=(x-mu)/sigma
    return norm.sf(z)

def generate_s_exp(n,beta):
    theta1=beta[0]
    theta2=beta[1]
    arr=np.zeros(n)
    for i in range(n):
        arr[i]=theta1*math.exp(-theta2*i/n)
    return arr

def generate_s_constant(n,mu):
    return np.repeat(mu,n)

def generate_s_powerlaw(n,beta):
    theta1 = beta[0]
    theta2 = beta[1]
    arr=np.zeros(n)
    for i in range(n):
        arr[i]=theta1*((1+(i+1)/n)**(-theta2))
    return arr

def generate_s_broken_powerlaw(n,beta,loc=0.5,slope2=0.5):
    mu = beta[0]
    slope1 = beta[1]
    breakpoint= int(n*loc)
    if breakpoint==0:
        return generate_s_powerlaw(n,np.array([mu,slope2]))
    arr=np.zeros(n)
    for i in range(breakpoint):
        arr[i]=mu*((1+(i+1)/n)**(-slope1))
    for i in range(n-breakpoint):
        arr[i]=arr[breakpoint-1]*((1+(i+1)/(n-breakpoint))**(-slope2))
    return arr

def generate_s_log_norm(n,mu,sigma):
    return np.exp(np.random.normal(np.log(mu),sigma,n))

def generate_s_gamma(n,alpha,beta):
    return gamma.rvs(alpha,loc=0,scale=1/beta,size=n)

def generate_s_unif(n,mu):
    return uniform.rvs(0,2*mu,size=n)

def generate_s(n,beta,snull):
    if snull=='exp':
        return generate_s_exp(n,beta)
    elif snull=='powerlaw':
        return generate_s_powerlaw(n,beta)
    elif snull=='constant':
        return generate_s_constant(n,beta[0])
    elif snull=='brokenpowerlaw':
        return generate_s_broken_powerlaw(n,beta) #break point is input as point/n.
    else:
        print("No this type of s!")
        return 0

def spectral_line(s,loc=0.75,strength=10,width=3):
    n=len(s)
    position=int(n*loc)
    for i in range(width):
        s[position+i]=strength
    return s

def generate_s_true(n,beta,strue,snull,loc=0.5,strength=10,width=5):
    if strue=='exp':
        return generate_s_exp(n,beta)
    elif strue=='powerlaw':
        return generate_s_powerlaw(n,beta)
    elif strue=='constant':
        return generate_s_constant(n,beta[0])
    elif strue=='lognorm':
        return generate_s_log_norm(n,beta[0],beta[1])
    elif strue=='gamma':
        return generate_s_gamma(n,beta[0],beta[1])
    elif strue=='brokenpowerlaw':
        return generate_s_broken_powerlaw(n,beta,loc=loc,slope2=strength) #beta=[mu,slope1,breakpoint,slope2] break point is input as point/n.
    elif strue=='unif':
        return generate_s_unif(n,beta[0])
    elif strue=='spectral_line':
        return spectral_line(generate_s(n,beta,snull),loc=loc,strength=strength,width=width)
    else:
        print("No this type of s!")
        return 0


def Max_mean_unit(mu):
    A=-0.56709
    B= -2.7336
    C= -2.3603
    D= 0.52816
    E= 0.33133
    F=1.0174
    alpha=3.9375
    beta= 0.48446
    return (A+B*mu+C*(mu-D)**2)*math.e**(-alpha*mu)+E*math.e**(-beta*mu)+F

def Max_mean(s):
    mean=0
    for x in s:
        mean+=Max_mean_unit(x)
    return mean

def Max_var_unit(mu):
    A=-3.1971
    B= 1.5118
    C= -1.5118
    D= 0.79384
    E= 1.9294
    F=6.1740
    G= 22.360/1000
    H= -7.2981
    I= 2.08378
    alpha= 0.750315
    beta=4.49654
    return (A+B*mu**2+C*(mu-D)**2)*math.e**(-alpha*mu)+(E+F*mu+G*(mu-H)**2)*math.e**(-beta*mu)+I

def Max_var(s):
    var=0
    for x in s:
        var+=Max_var_unit(x)
    return var

def poisson_dis(mu,i):
    return poisson.pmf(i,mu=mu)

def Sigma_diag(s,i,Q,n,max, epsilon):
    x=kapa12(s[i],max, epsilon)
    for j in range(n):
        x-=kapa11(s[j],max, epsilon)*Q[j,i]*kapa03(s[i],max, epsilon)
    return x

def expectation(s,n,X,I,max, epsilon):
    V = np.diag([s[i] for i in range(n)])
    Q = X * (X.T * V * X) ** (-1) * X.T
    Sigma = np.diag([Sigma_diag(s, i, Q, n,max, epsilon) for i in range(n)])
    E = -0.5*np.trace(X.T * Sigma * X * (X.T * V * X) ** (-1))
    for i in range(n):
        E += kapa1(s[i],max, epsilon)
    return float(E)

def uncon_expectation(s,n,X,I,max, epsilon):
    E=0
    for j in range(n):
        E += kapa1(s[j],max, epsilon)
    return float(E)

def uncon_var(s,n,X,I,max, epsilon):
    var=0
    for i in range(n):
        var += kapa2(s[i],max, epsilon)
    return var

def qtheta(s,n,X,I,max, epsilon):
    V = np.diag([s[i] for i in range(n)])
    k_11 = np.mat([kapa11(s[i], max, epsilon) for i in range(n)]).T
    return (k_11.T * X * (X.T * V * X) ** (-1) * X.T * k_11)[0, 0]

def Var(s,n,X,I,max, epsilon):
    return uncon_var(s,n,X,I,max, epsilon)-qtheta(s,n,X,I,max, epsilon)

def kapa1(mu,max,epsilon):
    x = 0
    for k in range(max):
        if k == 0:
            next_term=-2 * (k - mu) * poisson_dis(mu, k)
        else:
            next_term = 2 * (k * math.log(k / mu, math.e) - k + mu) * poisson_dis(mu, k)
        x += next_term
        if next_term<epsilon and k>mu:
            return x
    return x

def kapa2(mu,max, epsilon):
    x = 0
    k_1=kapa1(mu,max, epsilon)
    for k in range(max):
        if k == 0:
            next_term= (-2 * (k - mu) - k_1) ** 2 * poisson_dis(mu, k)
        else:
            next_term= (2 * (k * math.log(k / mu, math.e) - k + mu) - k_1) ** 2 * poisson_dis(mu, k)
        x += next_term
        if next_term<epsilon and k>mu:
            return x
    return x

def kapa11(mu,max, epsilon):
    x = 0
    k_1 = kapa1(mu, max, epsilon)
    for k in range(max):
        if k == 0:
            next_term= (-2 * (k - mu) - k_1) * (k - mu) * poisson_dis(mu, k)
        else:
            next_term= (2 * (k * math.log(k / mu, math.e) - k + mu) - k_1) * (k - mu) * poisson_dis(mu, k)
        x += next_term
        if next_term<epsilon and k>mu:
            return x
    return x

def kapa12(mu,max, epsilon):
    x = 0
    k_1 = kapa1(mu, max, epsilon)
    for k in range(max):
        if k == 0:
            next_term= (-2 * (k - mu) - k_1) * (k - mu) ** 2 * poisson_dis(mu, k)
        else:
            next_term= (2 * (k * math.log(k / mu, math.e) - k + mu) - k_1) * (k - mu) ** 2 * poisson_dis(mu,k)
        x += next_term
        if next_term<epsilon and k>mu:
            return x
    return x

def kapa03(mu,max,epsilon):
    return mu

def Cashstat(x,s):
    C=0.
    for j in range(len(x)):
        if x[j] <1e-5:
            C+=s[j]
        else:
            C +=s[j] - x[j] * np.log(s[j] / x[j]) - x[j]
    C=C*2
    return C

def p_value_chi(x,df): #Alg.1
    return chi2.sf(x,df)

def KMtest(x,s):  #Alg.2a
    return p_value_norm(x,Max_mean(s),math.sqrt(Max_var(s)))

def oracle_uncon_test(x,betatrue,n,snull,maximum=int(1e5),epsilon=1e-5):
    s = generate_s(n, betatrue, snull)
    I = np.mat([1.0 for i in range(len(betatrue))]).T
    X = design_mat(betatrue, n, snull)
    return p_value_norm(x, uncon_expectation(s, n, X, I, maximum, epsilon),
                        math.sqrt(uncon_var(s, n, X, I, maximum, epsilon)))

def uncon_plugin_test(x,beta,n,snull,maximum=int(1e5),epsilon=1e-5):
    s = generate_s(n, beta, snull)
    I = np.mat([1.0 for i in range(len(beta))]).T
    X = design_mat(beta, n, snull)
    return p_value_norm(x, uncon_expectation(s, n, X, I, maximum, epsilon),
                        math.sqrt(uncon_var(s, n, X, I, maximum, epsilon)))

def uncon_theory_test(x,beta,n,snull,maximum=int(1e5),epsilon=1e-5):  #Alg.3a
    s=generate_s(n,beta,snull)
    I=np.mat([1.0 for i in range(len(beta))]).T
    X=design_mat(beta,n,snull)
    return p_value_norm(x,uncon_expectation(s,n,X,I,maximum,epsilon),math.sqrt(Var(s,n,X,I,maximum,epsilon)))

def con_theory_test(x,beta,n,snull,maximum=int(1e5),epsilon=1e-5):  #Alg.3b
    s = generate_s(n, beta, snull)
    I = np.mat([1. for i in range(len(beta))]).T
    X = design_mat(beta, n, snull)
    return p_value_norm(x,expectation(s,n,X,I,maximum,epsilon),math.sqrt(Var(s,n,X,I,maximum,epsilon)))

def bootstrap_asymptotic(Cmin,beta,n,snull,B=1000, epsilon=1e-5):  #Alg.2b
    C = np.zeros(B)
    s=generate_s(n,beta,snull)
    for i in range(B):
        x = poisson_data(s)
        if np.all(np.abs(x) < 1e-5):
            continue
        bound = empirical_bounds(x, snull, epsilon)
        xopt = opt.minimize(LLF, beta, args=(x, snull), bounds=bound, method="L-BFGS-B")
        r = generate_s(n, xopt['x'], snull)
        C[i]=Cashstat(x,r)
    return p_value_norm(Cmin, statistics.mean(C), statistics.stdev(C))

def bootstrap_test(Cmin,beta,n,snull,B=1000, epsilon=1e-5):  #Alg.4a
    C = np.zeros(B)
    s=generate_s(n,beta,snull)
    for i in range(B):
        x = poisson_data(s)
        if np.all(np.abs(x) < 1e-5):
            continue
        bound = empirical_bounds(x, snull, epsilon)
        xopt = opt.minimize(LLF, beta, args=(x, snull), bounds=bound, method="L-BFGS-B")
        r = generate_s(n, xopt['x'], snull)
        C[i]=Cashstat(x,r)
    return np.mean(C >= Cmin)

def bootstrap_bias(Cmin,beta,n,snull,B1=1000,B2=1000, epsilon=1e-5):  #Alg.4b
    theta_tilde=np.zeros(len(beta))
    s = generate_s(n, beta, snull)
    for i in range(B1):
        x = poisson_data(s)
        if np.all(np.abs(x) < 1e-5):
            continue
        bound = empirical_bounds(x, snull, epsilon)
        xopt = opt.minimize(LLF, beta, args=(x, snull), bounds=bound, method="L-BFGS-B")
        theta_tilde += xopt['x']
    theta_adj=2*beta-theta_tilde/B1
    if theta_adj[0]<1e-5:
        theta_adj[0]=0
    return bootstrap_test(Cmin,theta_adj,n,snull,B2)

def double_boostrap(Cmin,beta,n,snull,B1=1000,B2=1000, epsilon=1e-5):  #Alg.4c
    C = np.zeros(B1)
    pvalue = np.zeros(B1)
    s = generate_s(n, beta, snull)
    for i in range(B1):
        x = poisson_data(s)
        if np.all(np.abs(x) < 1e-5):
            continue
        bound = empirical_bounds(x, snull, epsilon)
        xopt = opt.minimize(LLF, beta, args=(x, snull), bounds=bound, method="L-BFGS-B")
        r = generate_s(n, xopt['x'], snull)
        C[i] = Cashstat(x, r)
        pvalue[i]=bootstrap_test(C[i],xopt['x'],n,snull,B2)
    p = np.mean(C >= Cmin)
    return np.mean(pvalue <= p)

def LLF(theta,x,snull):
    if snull!='constant' and snull!='powerlaw' and snull!='exp':
        print("No such Likelihood Function!")
        return 0
    n = len(x)
    s = generate_s(n,theta,snull)
    return sum(s-x*np.log(s))

def design_mat(beta,n,snull):
    p=len(beta)
    if snull=='constant':
        return np.mat([1. for i in range(n)]).T
    elif snull=='powerlaw':
        return np.mat([[1 for x in range(n)], [math.log(1+(x + 1) / n, math.e) for x in range(n)]]).T
    elif snull=='exp':
        return np.mat([[1 for x in range(n)], [(x + 1) / n for x in range(n)]]).T
    else:
        print("Design Matrix Fail!")
        return 0

def single_test(n,B,B1,B2,beta,strue,snull,alpha,iters,maximum=int(1e5),epsilon=1e-5):
    s=generate_s_true(n,beta,strue,snull,loc=0.5,strength=10,width=5) # ground-truth. loc, strength & width are used for brokenpowerlaw, spectral line.
    rejections=np.zeros(10) #totally 8 methods, if strue=snull: Type I Error; if strue!=snull: Power
    for i in range(iters): # repetition
        print(f"Iteration {i + 1}")
        x=poisson_data(s)
        if np.all(np.abs(x) < 1e-5): # x==0, always accept
            continue
        bound = empirical_bounds(x, snull, epsilon)
        xopt = opt.minimize(LLF, beta, args=(x, snull), bounds=bound, method="L-BFGS-B") # Get MLE
        betahat=xopt['x']
        print(betahat)
        r = generate_s(n, betahat, snull) # null-distribution
        Cmin=Cashstat(x,r)

        # start test
        if p_value_chi(Cmin,n-len(betahat))<alpha: #Alg.1
            rejections[0]+=1
        if KMtest(Cmin,r)<alpha: #Alg.2a
            rejections[1]+=1
        if bootstrap_asymptotic(Cmin, betahat, n, snull, B)<alpha: #Alg.2b
            rejections[2]+=1
        if oracle_uncon_test(Cmin,beta, n, snull,maximum, epsilon)<alpha: rejections[3]+=1 #oracle uncon test, cannot be used when test power -- if H_0 not true, what is the true beta?
        if uncon_plugin_test(Cmin, betahat, n, snull, maximum, epsilon) < alpha: rejections[4] += 1  # plugin uncon test
        if uncon_theory_test(Cmin,betahat, n, snull,maximum, epsilon)<alpha: rejections[5]+=1 #Alg.3a
        if con_theory_test(Cmin,betahat, n, snull,maximum, epsilon)<alpha: rejections[6]+=1 #Alg.3b
        if bootstrap_test(Cmin,betahat, n, snull,B)<alpha: rejections[7]+=1 #4a
        if bootstrap_bias(Cmin,betahat,n,snull,B1,B2)<alpha: rejections[8]+=1 #4b
        if double_boostrap(Cmin,betahat,n,snull,B1,B2)<alpha: rejections[9]+=1 #4c

    return rejections/iters     #Rejection Rate

epsilon_error=1e-5 # error-control, user decide
maximum_iter=int(1e5) # maximum iterations, user decide

n=50  # number of bins
B=B1=B2=100  # bootstrap repetition times
beta=np.array([1,3])  #ground-truth beta*
strue='exp'  # true s
snull='exp'  # s of H_0
alpha=0.1  # significance level
iters=100  # repetition times
result=single_test(n, B, B1, B2, beta, strue, snull, alpha, iters,maximum_iter,epsilon_error)
print(result)
# Do not forget to store the result, like pd.to_cvs(...)

# You may add another loop for different n, beta, strue/snull and alpha
