import pandas as pd
import numpy as np
import math
import statistics
from scipy.stats import poisson, norm, chi2, gamma, uniform
import scipy.optimize as opt
from openpyxl import load_workbook

def write_to_exist_excel1(data,fileName, sheetName):
    df_old = pd.DataFrame(pd.read_excel(fileName, sheet_name=sheetName))  # 读取原数据文件和表
    row_old = df_old.shape[0]  # 获取原数据的行数

    df = data

    book = load_workbook(fileName)
    writer = pd.ExcelWriter(fileName, engine='openpyxl')
    writer._book = book
    writer._sheets = dict((ws.title, ws) for ws in book.worksheets)

    # 将data_added数据写入Excel中
    df.to_excel(writer, sheet_name=sheetName, startrow=row_old + 1, index=False, header=False)

    writer._save()  # 保存


# 将list[dict]类型的数据追加写入到现有的Excel中
def write_to_exist_excel2(data,fileName, sheetName):
    df_old = pd.DataFrame(pd.read_excel(fileName, sheet_name=sheetName))  # 读取原数据文件和表
    row_old = df_old.shape[0]  # 获取原数据的行数

    df = data

    book = load_workbook(fileName)
    writer = pd.ExcelWriter(fileName, engine='openpyxl')
    writer._book = book
    writer._sheets = dict((ws.title, ws) for ws in book.worksheets)

    # 将data_added数据写入Excel中
    df.to_excel(writer, sheet_name=sheetName, startrow=row_old + 1, index=False, header=False)

    writer._save()  # 保存


def normfun(x,mu,sigma):
    pdf=np.exp(-((x-mu)**2)/(2*sigma**2))/(sigma*np.sqrt(2*np.pi))
    return pdf

def poisson_data(mu):
    return np.random.poisson(mu)


def nb_data(r,mu):
    prob = r/(r+mu)
    return np.random.negative_binomial(r,prob)

def p_value_norm(x,mu,sigma):
    z=(x-mu)/sigma
    return norm.sf(z)

def generate_s_exp(n, beta):
    theta1, theta2 = beta
    i_over_n = np.arange(n) / n  # 生成 [0/n, 1/n, ..., (n-1)/n]
    return theta1 * np.exp(-theta2 * i_over_n)

def generate_s_constant(n,mu):
    return np.repeat(mu,n)

def generate_s_powerlaw(n, beta):
    mu, slope = beta
    i_values = np.arange(1, n+1)  # 生成 [1, 2, ..., n]
    return mu * (1 + i_values / n) ** (-slope)


def generate_s_broken_powerlaw(n, beta, loc=0.5, slope2=0.5):
    mu, slope1 = beta
    breakpoint = int(n * loc)
    if breakpoint == 0:
        return generate_s_powerlaw(n, np.array([mu, slope2]))

    # 前段（直接调用已优化的 generate_s_powerlaw）
    arr = np.empty(n)
    i_part1 = np.arange(1, breakpoint + 1)
    arr[:breakpoint] = mu * (1 + i_part1 / n) ** (-slope1)

    # 后段（基于前段最后一个值的向量化计算）
    i_part2 = np.arange(1, n - breakpoint + 1)
    decay_factor = (1 + i_part2 / (n - breakpoint)) ** (-slope2)
    arr[breakpoint:] = arr[breakpoint - 1] * decay_factor

    return arr

def generate_s_log_norm(n,mu,sigma):
    return np.exp(np.random.normal(np.log(mu),sigma,n))

def generate_s_gamma(n,alpha,beta):
    return gamma.rvs(alpha,loc=0,scale=1/beta,size=n)

def generate_s_unif(n,mu):
    return uniform.rvs(0,2*mu,size=n)

def generate_s(n,beta,snull,loc=0.5,strength=0.5,width=5):
    if snull=='exp':
        return generate_s_exp(n,beta)
    elif snull=='powerlaw':
        return generate_s_powerlaw(n,beta)
    elif snull=='constant':
        return generate_s_constant(n,beta[0])
    elif snull=='brokenpowerlaw':
        return generate_s_broken_powerlaw(n,beta,loc=loc,slope2=strength)
    else:
        print("No this type of s!")
        return 0

def spectral_line(s, loc=0.75, strength=10, width=3):
    position = int(len(s) * loc)
    end = min(position + width, len(s))
    s[position:end] = strength
    return s

def generate_s_true(n,beta,strue,snull,loc=0.5,strength=10,width=5):
    if strue=='exp':
        return generate_s_exp(n,beta)
    elif strue=='powerlaw':
        return generate_s_powerlaw(n,beta)
    elif strue=='constant':
        return generate_s_constant(n,beta[0])
    elif strue=='lognorm':
        return generate_s_log_norm(n,beta[0],beta[1])
    elif strue=='gamma':
        return generate_s_gamma(n,beta[0],beta[1])
    elif strue=='brokenpowerlaw':
        return generate_s_broken_powerlaw(n,beta,loc=loc,slope2=strength)
    elif strue=='unif':
        return generate_s_unif(n,beta[0])
    elif strue=='spectral_line':
        return spectral_line(generate_s(n,beta,snull),loc=loc,strength=strength,width=width)
    else:
        print("No this type of s!")
        return 0

def empirical_bounds(x,snull,epsilon=1e-5):
    bound = [[epsilon, 2*max(x)]] if snull == 'constant' else [[epsilon, 2*max(x)],
                                                             [-2*math.log((max(x) + 2 * epsilon) / (min(x) + epsilon), 2),
                                                              2*math.log((max(x) + 2 * epsilon) / (min(x) + epsilon), 2)]]
    return bound

def Cashstat(x,s):
    C=0.
    for j in range(len(x)):
        if x[j] <1e-5:
            C+=s[j]
        else:
            C +=s[j] - x[j] * np.log(s[j] / x[j]) - x[j]
    C=C*2
    return C